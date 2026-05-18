from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, TimeoutException

from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
from urllib.parse import urlparse
from pathlib import Path

import os
import re
import time
import random


# =========================
# 1. 运行配置
# =========================

BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"

# 综合模式输出
INTEGRATED_FILE_PATH = str(OUTPUT_DIR / "Raw_integrated_data.xlsx")

# 用户模式输出
USER_FILE_PATH = str(OUTPUT_DIR / "Raw_user_data.xlsx")
USER_VIDEO_FILE_PATH = str(OUTPUT_DIR / "Raw_user_video_data.xlsx")
USER_TEXT_FILE_PATH = str(OUTPUT_DIR / "Raw_user_text_data.xlsx")

# 小红书探索页入口
REDBOOK_PATH = "https://www.xiaohongshu.com/explore"

# ChromeDriver 配置：
# 1. 留空时，优先使用 Selenium Manager 自动处理驱动；
# 2. 若本机环境需要手动指定驱动，可填写本机 chromedriver.exe 路径。
CHROMEDRIVER_PATH = ""

# 浏览器选项
ENABLE_HEADLESS = False


# =========================
# 2. 通用工具函数
# =========================

def random_delay(base_delay=1, max_rand=2):
    """随机延迟，模拟人工操作。"""
    time.sleep(base_delay + random.uniform(0, max_rand))


def parse_count(raw_text):
    """
    将互动数文本转为整数。

    支持：
    1. 普通数字：123
    2. 带逗号数字：1,234
    3. 万级数字：1.2万
    """
    if not raw_text:
        return 0

    text = raw_text.strip().replace(",", "")

    try:
        if "万" in text:
            number = re.search(r"[\d.]+", text)
            return int(float(number.group()) * 10000) if number else 0

        number = re.search(r"\d+", text)
        return int(number.group()) if number else 0

    except Exception:
        return 0


def read_non_negative_int(prompt, allow_zero=True):
    """读取非负整数输入。"""
    while True:
        value = input(prompt).strip()
        if value.isdigit():
            num = int(value)
            if allow_zero or num > 0:
                return num
        print("请输入合法的整数。")


def read_choice(prompt, allowed_choices):
    """读取限定选项输入。"""
    allowed_choices = {str(choice) for choice in allowed_choices}

    while True:
        value = input(prompt).strip()
        if value in allowed_choices:
            return value
        print(f"请输入以下选项之一：{', '.join(sorted(allowed_choices))}")


def split_keywords(raw_keywords):
    """将空格分隔的多个关键词拆分为列表。"""
    return [item for item in raw_keywords.split() if item.strip()]


def sanitize_sheet_name(sheet_name, fallback="data"):
    """生成合法的 Excel 工作表名称。"""
    if not sheet_name:
        sheet_name = fallback

    safe_name = re.sub(r'[:\\/?*\[\]]', "_", sheet_name).strip()
    safe_name = safe_name[:31] if safe_name else fallback
    return safe_name or fallback


def create_driver():
    """创建 Chrome WebDriver。"""
    options = Options()

    if ENABLE_HEADLESS:
        options.add_argument("--headless=new")

    if CHROMEDRIVER_PATH.strip():
        service = ChromeService(executable_path=CHROMEDRIVER_PATH.strip())
        return webdriver.Chrome(service=service, options=options)

    return webdriver.Chrome(options=options)


# =========================
# 3. 时间与链接处理工具类
# =========================

class OtherDataDeal:
    """时间解析、笔记 ID 提取。"""

    def DateDeal(self, raw_date):
        """将相对时间解析为 YYYY-MM-DD。"""
        if not raw_date:
            return "无"

        relative_text = raw_date.strip().replace("周", "星期")
        today = datetime.now().date()
        now = datetime.now()

        week_map = {
            "星期一": 0,
            "星期二": 1,
            "星期三": 2,
            "星期四": 3,
            "星期五": 4,
            "星期六": 5,
            "星期日": 6
        }

        try:
            relative_text = relative_text.replace("编辑于", "").strip()
            core_text = relative_text.rsplit(" ", 1)[0] if " " in relative_text else relative_text

            if "刚刚" in core_text or "分钟前" in core_text:
                return today.strftime("%Y-%m-%d")

            if "昨天" in core_text:
                return (today - timedelta(days=1)).strftime("%Y-%m-%d")

            day_match = re.search(r"(\d+)天前", core_text)
            if day_match:
                days_ago = int(day_match.group(1))
                return (today - timedelta(days=days_ago)).strftime("%Y-%m-%d")

            hour_match = re.search(r"(\d+)小时前", core_text)
            if hour_match:
                hours_ago = int(hour_match.group(1))
                return (now - timedelta(hours=hours_ago)).date().strftime("%Y-%m-%d")

            for week_text, week_num in week_map.items():
                if week_text in core_text:
                    today_week_num = today.weekday()
                    days_to_subtract = (today_week_num - week_num) % 7
                    return (today - timedelta(days=days_to_subtract)).strftime("%Y-%m-%d")

            if re.match(r"^\d{1,2}-\d{1,2}$", core_text):
                current_year = datetime.now().year
                month, day = map(int, core_text.split("-"))
                guessed_date = datetime(current_year, month, day).date()

                if guessed_date > today:
                    guessed_date = datetime(current_year - 1, month, day).date()

                return guessed_date.strftime("%Y-%m-%d")

            return core_text

        except Exception as e:
            print(f"解析时间失败：{raw_date}，错误：{e}")
            return raw_date

    @staticmethod
    def extract_note_id(note_url):
        """从笔记 URL 中提取笔记 ID。"""
        try:
            parsed = urlparse(note_url)
            path_parts = [part for part in parsed.path.split("/") if part]

            # 常见形式：
            # /explore/xxxx
            if len(path_parts) >= 2 and path_parts[0] == "explore":
                return path_parts[1]

            # /discovery/item/xxxx
            if len(path_parts) >= 3 and path_parts[0] == "discovery" and path_parts[1] == "item":
                return path_parts[2]

            # /user/profile/用户ID/笔记ID
            if len(path_parts) >= 4 and path_parts[0] == "user" and path_parts[1] == "profile":
                return path_parts[3]

            # 兜底：取最后一个路径段
            return path_parts[-1] if path_parts else ""

        except Exception as e:
            print(f"提取笔记 ID 失败：{note_url}，错误：{e}")
            return ""


# =========================
# 4. Excel 写入工具类
# =========================

class ExcelWriter:
    """Excel 数据写入工具。"""

    @staticmethod
    def WriteDataToExcel(data_list, file_path, sheet_name):
        if not data_list:
            return

        try:
            sheet_name = sanitize_sheet_name(sheet_name)

            folder = os.path.dirname(file_path)
            if folder and not os.path.exists(folder):
                os.makedirs(folder, exist_ok=True)

            if not os.path.exists(file_path):
                wb = Workbook()
                ws = wb.active
                ws.title = sheet_name
                wb.save(file_path)
                wb.close()

            wb = load_workbook(file_path)

            if sheet_name not in wb.sheetnames:
                ws = wb.create_sheet(sheet_name)
            else:
                ws = wb[sheet_name]

            current_headers = []
            if ws.max_row >= 1:
                current_headers = [cell.value for cell in ws[1] if cell.value]

            all_new_keys = []
            for row in data_list:
                for key in row.keys():
                    if key not in all_new_keys:
                        all_new_keys.append(key)

            if not current_headers:
                current_headers = all_new_keys
                ws.append(current_headers)
            else:
                missing_headers = [key for key in all_new_keys if key not in current_headers]
                if missing_headers:
                    for header in missing_headers:
                        current_headers.append(header)
                        ws.cell(row=1, column=len(current_headers), value=header)

            for data_row in data_list:
                row_values = [data_row.get(header, "") for header in current_headers]
                ws.append(row_values)

            wb.save(file_path)
            wb.close()

            print(f"已写入 {len(data_list)} 条数据到：{file_path}，工作表：{sheet_name}")

        except Exception as e:
            print(f"写入 Excel 失败：{e}")


# =========================
# 5. 笔记数据爬取类
# =========================

class NoteDataSpider:
    """负责笔记、评论、综合模式笔记和用户主页笔记的提取。"""

    def __init__(self):
        self.other_deal = OtherDataDeal()

    def Comment(self, driver, comment_count_to_crawl):
        """爬取指定数量的评论数据。"""
        if comment_count_to_crawl <= 0:
            return []

        comments = []

        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[@class="content"] | //div[@class="interactions"]')
                )
            )

            content_items = driver.find_elements(By.CSS_SELECTOR, "div.content")
            date_items = driver.find_elements(By.CSS_SELECTOR, "div.date")
            interactions_items = driver.find_elements(By.CSS_SELECTOR, "div.interactions")

            actual_comment_count = min(
                len(content_items),
                len(date_items),
                len(interactions_items),
                comment_count_to_crawl
            )

            for idx in range(actual_comment_count):
                comment_data = {
                    "content": "无",
                    "publish_time": "无",
                    "location": "无",
                    "like_count": 0,
                    "reply_count": 0
                }

                try:
                    content_elem = content_items[idx].find_element(By.CSS_SELECTOR, "span.note-text")
                    content_text = content_elem.text.strip()
                    comment_data["content"] = content_text if content_text else "无"
                except Exception:
                    pass

                try:
                    date_item = date_items[idx]
                    full_text = date_item.text.strip()

                    loc_text = ""
                    try:
                        loc_elem = date_item.find_element(By.CSS_SELECTOR, "span.location")
                        loc_text = loc_elem.text.strip()
                    except Exception:
                        pass

                    time_text = full_text.replace(loc_text, "").strip() if loc_text else full_text
                    comment_data["publish_time"] = (
                        self.other_deal.DateDeal(time_text) if time_text else "无"
                    )
                    comment_data["location"] = loc_text if loc_text else "无"

                except Exception:
                    pass

                try:
                    interaction_text = interactions_items[idx].text.strip()
                    numbers = re.findall(r"[\d.]+万?|\d+", interaction_text)

                    if len(numbers) >= 1:
                        comment_data["like_count"] = parse_count(numbers[0])

                    if len(numbers) >= 2:
                        comment_data["reply_count"] = parse_count(numbers[1])

                except Exception:
                    pass

                comments.append(comment_data)

        except TimeoutException:
            print("未找到评论区域")
        except Exception as e:
            print(f"获取评论失败：{e}")

        while len(comments) < comment_count_to_crawl:
            comments.append({
                "content": "无",
                "publish_time": "无",
                "location": "无",
                "like_count": 0,
                "reply_count": 0
            })

        return comments

    def GetNoteDate(self, driver):
        """提取当前打开笔记的发布时间。"""
        pure_date = ""

        try:
            date_element = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[contains(@class, "bottom-container")]//span[@class="date"]')
                )
            )
            raw_text = date_element.text.strip()
            pure_date = self.other_deal.DateDeal(raw_text)

        except TimeoutException:
            pass
        except Exception as e:
            print(f"提取日期出现错误：{e}")

        return pure_date

    def NoteCommon(self, driver, note_href, comment_count_to_crawl):
        """提取视频和图文笔记的公共字段。"""
        common_data = {
            "note_id": OtherDataDeal.extract_note_id(note_href),
            "name": "",
            "tags": "",
            "publish_time": "",
            "like_count": 0,
            "collect_count": 0,
            "comment_count": 0,
            "note_url": note_href
        }

        try:
            comments = self.Comment(driver, comment_count_to_crawl)

            try:
                title_elem = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.ID, "detail-title"))
                )
                common_data["name"] = title_elem.text.strip()
            except Exception:
                pass

            try:
                tag_elems = driver.find_elements(By.CSS_SELECTOR, "a.tag")
                tags = [
                    tag.text.strip()
                    for tag in tag_elems
                    if tag.text.strip().startswith("#")
                ]
                common_data["tags"] = ",".join(tags)
            except Exception:
                pass

            try:
                date_elem = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[contains(@class, "bottom-container")]//span[@class="date"]')
                    )
                )
                raw_time = date_elem.text.strip()
                common_data["publish_time"] = self.other_deal.DateDeal(raw_time)
            except Exception:
                pass

            try:
                interact_elems = driver.find_elements(
                    By.CSS_SELECTOR,
                    "div.interact-container .buttons .count"
                )

                if len(interact_elems) >= 3:
                    common_data["like_count"] = parse_count(interact_elems[0].text)
                    common_data["collect_count"] = parse_count(interact_elems[1].text)
                    common_data["comment_count"] = parse_count(interact_elems[2].text)

            except Exception as e:
                print(f"获取互动数据失败：{e}")

            for idx, comment in enumerate(comments, 1):
                common_data[f"comment{idx}_content"] = comment["content"]
                common_data[f"comment{idx}_time"] = comment["publish_time"]
                common_data[f"comment{idx}_loc"] = comment["location"]
                common_data[f"comment{idx}_like"] = comment["like_count"]
                common_data[f"comment{idx}_reply"] = comment["reply_count"]

        except Exception as e:
            print(f"获取笔记公共数据失败：{e}")

        return common_data

    def VideoNote(self, driver, note_href, comment_count_to_crawl):
        """提取视频笔记数据。"""
        video_data = self.NoteCommon(driver, note_href, comment_count_to_crawl)
        video_data["note_type"] = "video"

        try:
            ActionChains(driver).send_keys(Keys.SPACE).perform()
            random_delay(1, 1)
        except Exception:
            pass

        try:
            total_time_elem = driver.find_element(By.CSS_SELECTOR, "xg-time span:nth-child(2)")
            video_data["video_length"] = total_time_elem.text.strip()
        except Exception:
            video_data["video_length"] = "00:00"

        print("-" * 35 + " 提取到的视频笔记数据 " + "-" * 35)
        for key, value in video_data.items():
            print(f"{key}: {value}")
        print("-" * 100)

        return video_data

    def TextNote(self, driver, note_href, comment_count_to_crawl):
        """提取图文笔记数据。"""
        text_data = self.NoteCommon(driver, note_href, comment_count_to_crawl)
        text_data["note_type"] = "text"

        try:
            content_elem = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="detail-desc"]//span'))
            )
            content = content_elem.text.strip().replace("\n", " ").replace("\xa0", " ")
            text_data["content"] = content if content else "无"
        except Exception as e:
            print(f"获取图文笔记内容失败：{e}")
            text_data["content"] = "无"

        print("-" * 35 + " 提取到的图文笔记数据 " + "-" * 35)
        for key, value in text_data.items():
            print(f"{key}: {value}")
        print("-" * 100)

        return text_data

    def _get_new_note_snapshots(self, driver, visited_note_urls):
        """
        获取当前页面中未处理过的笔记快照。
        只保存 href 和类型，不保存元素对象，避免返回页面后元素失效。
        """
        snapshots = []
        note_items = driver.find_elements(By.CSS_SELECTOR, "section.note-item")

        for item in note_items:
            try:
                link_elem = item.find_element(By.CSS_SELECTOR, "a.cover.mask.ld")
                note_href = link_elem.get_attribute("href")

                if not note_href or note_href in visited_note_urls:
                    continue

                is_video = False
                try:
                    item.find_element(By.CSS_SELECTOR, "span.play-icon")
                    is_video = True
                except NoSuchElementException:
                    pass

                snapshots.append({
                    "href": note_href,
                    "is_video": is_video
                })

            except NoSuchElementException:
                continue
            except Exception:
                continue

        return snapshots

    def _open_note_by_href(self, driver, note_href):
        """
        根据 href 重新定位当前页面中的笔记并点击。
        避免使用旧元素对象导致页面返回后元素失效。
        """
        note_items = driver.find_elements(By.CSS_SELECTOR, "section.note-item")

        for item in note_items:
            try:
                link_elem = item.find_element(By.CSS_SELECTOR, "a.cover.mask.ld")
                current_href = link_elem.get_attribute("href")

                if current_href == note_href:
                    driver.execute_script(
                        "arguments[0].scrollIntoView({block: 'center'});",
                        item
                    )
                    random_delay(1, 1)

                    try:
                        link_elem.click()
                    except Exception:
                        driver.execute_script("arguments[0].click();", link_elem)

                    return True

            except Exception:
                continue

        return False

    def ClickLastNote(self, driver):
        """点击当前页面列表中最后一条笔记，用于获取最早笔记发布时间。"""
        try:
            all_notes = WebDriverWait(driver, 10).until(
                EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.note-item"))
            )

            if not all_notes:
                print("页面中未找到任何笔记")
                return False

            last_note = all_notes[-1]
            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", last_note)
            random_delay(1, 1)

            try:
                note_click_area = last_note.find_element(By.CSS_SELECTOR, "a.cover.mask.ld")
                note_click_area.click()
            except Exception:
                last_note.click()

            return True

        except TimeoutException:
            return False
        except Exception as e:
            print(f"点击最后一个笔记出现错误：{e}")
            return False

    def CollectIntegrateNotes(
        self,
        driver,
        target_total_count,
        target_video_count,
        target_text_count,
        comment_count_to_crawl
    ):
        """
        综合模式收集笔记。

        两种方式：
        1. target_total_count > 0：按总数量爬取；
        2. target_total_count == 0：分别按视频数量、图文数量爬取。
        """
        video_notes_data = []
        text_notes_data = []

        collected_video_count = 0
        collected_text_count = 0

        visited_note_urls = set()

        last_scroll_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts_without_new_content = 5

        is_total_mode = target_total_count > 0

        def should_stop():
            if is_total_mode:
                return collected_video_count + collected_text_count >= target_total_count
            return (
                collected_video_count >= target_video_count
                and collected_text_count >= target_text_count
            )

        def need_this_type(is_video):
            if is_total_mode:
                return True

            if is_video:
                return collected_video_count < target_video_count

            return collected_text_count < target_text_count

        while not should_stop():
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.note-item"))
                )
            except TimeoutException:
                print("未找到搜索结果中的笔记卡片，停止本轮爬取")
                break

            snapshots = self._get_new_note_snapshots(driver, visited_note_urls)

            for note_snapshot in snapshots:
                if should_stop():
                    break

                note_href = note_snapshot["href"]
                is_video = note_snapshot["is_video"]

                if not need_this_type(is_video):
                    visited_note_urls.add(note_href)
                    continue

                visited_note_urls.add(note_href)

                try:
                    opened = self._open_note_by_href(driver, note_href)

                    if not opened:
                        print(f"未能打开笔记：{note_href}")
                        continue

                    random_delay(3, 4)

                    if is_video:
                        note_data = self.VideoNote(
                            driver,
                            note_href,
                            comment_count_to_crawl
                        )
                        video_notes_data.append(note_data)
                        collected_video_count += 1
                        print(f"已收集视频笔记 {collected_video_count} 条")
                    else:
                        note_data = self.TextNote(
                            driver,
                            note_href,
                            comment_count_to_crawl
                        )
                        text_notes_data.append(note_data)
                        collected_text_count += 1
                        print(f"已收集图文笔记 {collected_text_count} 条")

                    driver.back()
                    random_delay(2, 2)

                    WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.note-item"))
                    )

                except Exception as e:
                    print(f"处理单条笔记时出错：{e}")

                    try:
                        driver.back()
                        random_delay(2, 2)
                    except Exception:
                        pass

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            random_delay(2, 3)

            new_scroll_height = driver.execute_script("return document.body.scrollHeight")

            if new_scroll_height == last_scroll_height:
                scroll_attempts += 1

                if scroll_attempts >= max_scroll_attempts_without_new_content:
                    print("页面已多次滚动但没有新内容加载，停止爬取")
                    break
            else:
                scroll_attempts = 0
                last_scroll_height = new_scroll_height

        final_video_count = len(video_notes_data)
        final_text_count = len(text_notes_data)
        final_total_count = final_video_count + final_text_count

        print(
            f"综合模式爬取完成："
            f"视频笔记 {final_video_count} 条，"
            f"图文笔记 {final_text_count} 条，"
            f"总计 {final_total_count} 条"
        )

        return (
            video_notes_data,
            text_notes_data,
            final_video_count,
            final_text_count,
            final_total_count
        )

    def CollectUserHomeNotes(
        self,
        driver,
        target_video_count,
        target_text_count,
        comment_count_to_crawl,
        user_id
    ):
        """在单个用户主页中采集指定数量的视频笔记和图文笔记。"""
        video_notes_data = []
        text_notes_data = []

        collected_video_count = 0
        collected_text_count = 0

        visited_note_urls = set()
        last_scroll_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts_without_new_content = 5

        def should_stop():
            return (
                collected_video_count >= target_video_count
                and collected_text_count >= target_text_count
            )

        def need_this_type(is_video):
            if is_video:
                return collected_video_count < target_video_count
            return collected_text_count < target_text_count

        if should_stop():
            return video_notes_data, text_notes_data

        while not should_stop():
            try:
                WebDriverWait(driver, 8).until(
                    EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.note-item"))
                )
            except TimeoutException:
                print("用户主页未找到可采集笔记，停止笔记采集")
                break

            snapshots = self._get_new_note_snapshots(driver, visited_note_urls)

            for note_snapshot in snapshots:
                if should_stop():
                    break

                note_href = note_snapshot["href"]
                is_video = note_snapshot["is_video"]

                if not need_this_type(is_video):
                    visited_note_urls.add(note_href)
                    continue

                visited_note_urls.add(note_href)

                try:
                    opened = self._open_note_by_href(driver, note_href)

                    if not opened:
                        print(f"未能打开用户主页笔记：{note_href}")
                        continue

                    random_delay(3, 4)

                    if is_video:
                        note_data = self.VideoNote(
                            driver,
                            note_href,
                            comment_count_to_crawl
                        )
                        note_data["user_id"] = user_id
                        video_notes_data.append(note_data)
                        collected_video_count += 1
                        print(f"已收集用户视频笔记 {collected_video_count}/{target_video_count} 条")
                    else:
                        note_data = self.TextNote(
                            driver,
                            note_href,
                            comment_count_to_crawl
                        )
                        note_data["user_id"] = user_id
                        text_notes_data.append(note_data)
                        collected_text_count += 1
                        print(f"已收集用户图文笔记 {collected_text_count}/{target_text_count} 条")

                    driver.back()
                    random_delay(2, 2)

                    WebDriverWait(driver, 10).until(
                        EC.presence_of_all_elements_located((By.CSS_SELECTOR, "section.note-item"))
                    )

                except Exception as e:
                    print(f"处理用户主页单条笔记时出错：{e}")

                    try:
                        driver.back()
                        random_delay(2, 2)
                    except Exception:
                        pass

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            random_delay(2, 3)

            new_scroll_height = driver.execute_script("return document.body.scrollHeight")

            if new_scroll_height == last_scroll_height:
                scroll_attempts += 1

                if scroll_attempts >= max_scroll_attempts_without_new_content:
                    print("用户主页多次滚动后没有新内容，停止笔记采集")
                    break
            else:
                scroll_attempts = 0
                last_scroll_height = new_scroll_height

        print(
            f"用户主页笔记采集完成："
            f"视频笔记 {len(video_notes_data)} 条，"
            f"图文笔记 {len(text_notes_data)} 条"
        )

        return video_notes_data, text_notes_data


# =========================
# 6. 综合模式主爬虫类
# =========================

class RedBookIntegratedSpider:
    """综合搜索结果页笔记采集。"""

    def __init__(self):
        self.note_spider = NoteDataSpider()
        self.excel_writer = ExcelWriter()

    def Search(self, driver, search_query):
        """执行搜索。"""
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "search-input"))
        )

        search_input.clear()
        search_input.send_keys(search_query)
        search_input.send_keys(Keys.ENTER)

        print(f"已执行搜索：{search_query}")
        random_delay(3, 4)

    def SelectAllCategory(self, driver):
        """选择综合搜索结果。"""
        try:
            all_category = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//div[@class="content-container"]//div[@id="all"]')
                )
            )
            all_category.click()
            print("已选择综合模式")
            random_delay(2, 2)

        except Exception as e:
            print(f"点击综合模式选项失败，尝试继续执行：{e}")

    def ClearSearchInput(self, driver):
        """清空搜索框。"""
        try:
            search_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "search-input"))
            )
            search_input.click()
            search_input.send_keys(Keys.CONTROL, "a")
            search_input.send_keys(Keys.BACKSPACE)
            random_delay(1, 1)
        except Exception as e:
            print(f"清空搜索框失败：{e}")

    def WriteIntegrateAnalysis(self, search_query, total_count, video_count, text_count):
        """写入综合统计分析。"""
        video_ratio = (video_count / total_count * 100) if total_count > 0 else 0.0
        text_ratio = (text_count / total_count * 100) if total_count > 0 else 0.0

        video_text_ratio = (video_count / text_count) if text_count > 0 else "无穷大"
        text_video_ratio = (text_count / video_count) if video_count > 0 else "无穷大"

        analysis_data = {
            "搜索关键词": search_query,
            "总爬取数量": total_count,
            "视频笔记数量": video_count,
            "图文笔记数量": text_count,
            "视频笔记占比(%)": round(video_ratio, 2),
            "图文笔记占比(%)": round(text_ratio, 2),
            "视频-图文相对比例": round(video_text_ratio, 2) if isinstance(video_text_ratio, (int, float)) else video_text_ratio,
            "图文-视频相对比例": round(text_video_ratio, 2) if isinstance(text_video_ratio, (int, float)) else text_video_ratio,
            "爬取时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }

        self.excel_writer.WriteDataToExcel(
            [analysis_data],
            INTEGRATED_FILE_PATH,
            "integrated_analysis"
        )

    def HandleIntegrateMode(
        self,
        driver,
        search_query,
        target_total_count,
        target_video_count,
        target_text_count,
        comment_count_to_crawl
    ):
        """执行综合模式完整流程。"""
        (
            video_notes,
            text_notes,
            video_count,
            text_count,
            total_count
        ) = self.note_spider.CollectIntegrateNotes(
            driver,
            target_total_count,
            target_video_count,
            target_text_count,
            comment_count_to_crawl
        )

        all_notes = video_notes + text_notes

        if all_notes:
            self.excel_writer.WriteDataToExcel(
                all_notes,
                INTEGRATED_FILE_PATH,
                "integrated_note"
            )

        self.WriteIntegrateAnalysis(
            search_query,
            total_count,
            video_count,
            text_count
        )


# =========================
# 7. 用户模式主爬虫类
# =========================

class RedBookUserSpider:
    """用户搜索结果页与用户主页数据采集。"""

    def __init__(self):
        self.note_spider = NoteDataSpider()
        self.excel_writer = ExcelWriter()

    def Search(self, driver, search_query):
        """执行搜索。"""
        search_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "search-input"))
        )

        search_input.clear()
        search_input.send_keys(search_query)
        search_input.send_keys(Keys.ENTER)

        print(f"已执行搜索：{search_query}")
        random_delay(3, 4)

    def SelectUserCategory(self, driver):
        """选择用户搜索结果。"""
        try:
            user_category = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//div[@class="content-container"]//div[@id="user"]')
                )
            )
            user_category.click()
            print("已选择用户模式")
            random_delay(2, 2)

        except Exception as e:
            print(f"点击用户模式选项失败，尝试继续执行：{e}")

    def ClearSearchInput(self, driver):
        """清空搜索框。"""
        try:
            search_input = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.ID, "search-input"))
            )
            search_input.click()
            search_input.send_keys(Keys.CONTROL, "a")
            search_input.send_keys(Keys.BACKSPACE)
            random_delay(1, 1)
        except Exception as e:
            print(f"清空搜索框失败：{e}")

    def UserDataGet(self, driver, user_url):
        """提取单个用户主页的基础数据。"""
        user_data = {
            "User ID": "",
            "Account Name": "",
            "Certification Info": "",
            "Creation Time": "",
            "Gender": "",
            "Age": "",
            "Tags": "",
            "IP Location": "",
            "Following": "",
            "Followers": "",
            "Likes and Collections": "",
            "Video Count": "",
            "Text Count": "",
            "User URL": user_url,
            "Bio": ""
        }

        try:
            try:
                account_name_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[@class="user-name"] | //h1[contains(@class, "user-name")]')
                    )
                )
                user_data["Account Name"] = account_name_element.text.strip()
            except Exception:
                pass

            try:
                WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[@class="user-nickname"]//span[contains(@class, "verify-icon")]')
                    )
                )
                user_data["Certification Info"] = 1
            except TimeoutException:
                user_data["Certification Info"] = 0
            except Exception as e:
                print(f"爬取认证信息错误：{e}")

            try:
                red_id_element = WebDriverWait(driver, 2).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//span[contains(text(), "小红书号：")]')
                    )
                )
                red_id_text = red_id_element.text
                user_id_match = re.search(r"小红书号[\s:：]+(.+)", red_id_text)
                user_data["User ID"] = user_id_match.group(1) if user_id_match else red_id_text
            except Exception as e:
                print(f"爬取账号 ID 错误：{e}")

            try:
                ip_span = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//span[contains(@class, "user-IP")]')
                    )
                )
                parent = ip_span.find_element(By.XPATH, "ancestor::div[1]")
                text = driver.execute_script("return arguments[0].innerText;", parent)
                match = re.search(r"IP\s*属地[:：]\s*([^\s]+)", text)
                user_data["IP Location"] = match.group(1) if match else 0
            except TimeoutException:
                user_data["IP Location"] = 0
            except Exception as e:
                print(f"IP 归属地提取失败：{e}")
                user_data["IP Location"] = 0

            try:
                bio_element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.XPATH, '//div[contains(@class, "user-desc")]'))
                )
                user_data["Bio"] = bio_element.text.strip()
            except TimeoutException:
                user_data["Bio"] = "无"
            except Exception as e:
                print(f"提取简介时发生错误：{e}")

            user_data["Gender"] = 0
            user_data["Age"] = 0
            other_tags = []

            try:
                gender_container = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//div[contains(@class, "gender")]')
                    )
                )

                gender_html = gender_container.get_attribute("innerHTML").lower()
                if "female" in gender_html:
                    gender = "F"
                elif "male" in gender_html:
                    gender = "M"
                else:
                    gender = 0

                user_data["Gender"] = gender

                try:
                    age_text = gender_container.find_element(
                        By.XPATH, './/span[contains(@class, "gender-text")]'
                    ).text.strip()
                    age_match = re.search(r"(\d+)岁", age_text)
                    if age_match:
                        user_data["Age"] = age_match.group(1)
                except NoSuchElementException:
                    pass

            except TimeoutException:
                user_data["Gender"] = 0
                user_data["Age"] = 0
            except Exception as e:
                print(f"性别/年龄解析失败：{e}")

            try:
                tag_items = WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, '//div[contains(@class, "user-tags")]/div[contains(@class, "tag-item")]')
                    )
                )
                for tag_item in tag_items:
                    try:
                        tag_text = tag_item.text.strip()
                        if tag_text and not tag_text.isdigit():
                            other_tags.append(tag_text)
                    except Exception as e:
                        print(f"解析标签失败：{e}")
                user_data["Tags"] = ", ".join(other_tags) if other_tags else ""
            except TimeoutException:
                user_data["Tags"] = ""
            except Exception as e:
                print(f"提取标签时出错：{e}")

            try:
                interaction_items = WebDriverWait(driver, 5).until(
                    EC.presence_of_all_elements_located(
                        (By.XPATH, '//div[contains(@class, "user-interactions")]/div')
                    )
                )
                for item in interaction_items:
                    count_text = item.find_element(By.CLASS_NAME, "count").text.strip()
                    shows_text = item.find_element(By.CLASS_NAME, "shows").text.strip()
                    if "关注" in shows_text:
                        user_data["Following"] = count_text
                    elif "粉丝" in shows_text:
                        user_data["Followers"] = count_text
                    elif "获赞与收藏" in shows_text:
                        user_data["Likes and Collections"] = count_text
            except Exception as e:
                print(f"提取关注数、粉丝数、获赞与收藏数时发生错误：{e}")

        except Exception as e:
            print(f"解析用户 {user_url} 资料时发生错误：{e}")

        return user_data

    def scroll_to_bottom(self, driver):
        """滚动到页面底部，直到无法加载更多内容。"""
        last_height = driver.execute_script("return document.body.scrollHeight")
        attempts = 0
        max_attempts = 10

        while attempts < max_attempts:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            random_delay(2, 1)

            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                attempts += 1
            else:
                attempts = 0
                last_height = new_height

    def collect_user_links(self, driver, target_user_count):
        """从用户搜索结果页收集用户主页链接。"""
        print("获取用户列表…")

        if target_user_count <= 0:
            print("请输入一个正整数")
            return []

        visited_user_urls = set()
        user_links_to_visit = []

        last_scroll_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempts = 0
        max_scroll_attempts_without_new_content = 5

        print(f"正在加载用户列表，目标数量：{target_user_count}")

        while (
            len(user_links_to_visit) < target_user_count
            and scroll_attempts < max_scroll_attempts_without_new_content
        ):
            current_page_user_items = driver.find_elements(By.CLASS_NAME, "user-list-item")

            for item in current_page_user_items:
                try:
                    link_element = item.find_element(By.TAG_NAME, "a")
                    user_url = link_element.get_attribute("href")

                    if user_url and user_url not in visited_user_urls:
                        user_links_to_visit.append(user_url)
                        visited_user_urls.add(user_url)

                        if len(user_links_to_visit) >= target_user_count:
                            break

                except Exception:
                    continue

            if len(user_links_to_visit) >= target_user_count:
                break

            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            random_delay(2, 1)

            new_scroll_height = driver.execute_script("return document.body.scrollHeight")
            if new_scroll_height == last_scroll_height:
                scroll_attempts += 1
            else:
                scroll_attempts = 0
                last_scroll_height = new_scroll_height

        print(f"共收集到 {len(user_links_to_visit)} 个待访问的用户链接")
        return user_links_to_visit

    def count_loaded_notes(self, driver):
        """统计当前用户主页已加载区域中的视频笔记数和图文笔记数。"""
        video_count = 0
        text_count = 0

        try:
            all_notes_on_page = driver.find_elements(By.CSS_SELECTOR, "section.note-item")

            for note_elem in all_notes_on_page:
                try:
                    note_elem.find_element(By.CSS_SELECTOR, "span.play-icon")
                    video_count += 1
                except NoSuchElementException:
                    text_count += 1

        except Exception as e:
            print(f"统计主页笔记数量失败：{e}")

        return video_count, text_count

    def get_creation_time_from_last_note(self, driver):
        """尝试通过最早一条可见笔记的发布时间估计账号内容创建时间。"""
        creation_time = ""

        try:
            last_note_clicked = self.note_spider.ClickLastNote(driver)

            if last_note_clicked:
                random_delay(3, 2)
                creation_time = self.note_spider.GetNoteDate(driver)

                try:
                    ActionChains(driver).key_down(Keys.ESCAPE).key_up(Keys.ESCAPE).perform()
                    random_delay(1, 0.5)
                except Exception as e:
                    print(f"关闭笔记详情失败：{e}，尝试后退返回")
                    driver.back()
                    random_delay(2, 1)

        except Exception as e:
            print(f"获取创建时间失败：{e}")

        return creation_time

    def CollectUsersBySearchResults(
        self,
        driver,
        search_query,
        target_user_count,
        target_video_count,
        target_text_count,
        comment_count_to_crawl
    ):
        """访问用户搜索结果中的主页，并采集用户资料及其笔记。"""
        initial_list_page_url = driver.current_url
        sheet_name = sanitize_sheet_name(search_query, fallback="user_search")

        user_links_to_visit = self.collect_user_links(driver, target_user_count)

        for index, user_url in enumerate(user_links_to_visit, 1):
            print(f"\n[{index}/{len(user_links_to_visit)}] 正在访问用户主页：{user_url}")

            try:
                driver.get(user_url)
                random_delay(5, 3)
                print(f"已访问第 {index} 个用户主页")

                user_data = self.UserDataGet(driver, user_url)
                user_id = user_data.get("User ID", "")

                video_notes, text_notes = self.note_spider.CollectUserHomeNotes(
                    driver,
                    target_video_count,
                    target_text_count,
                    comment_count_to_crawl,
                    user_id
                )

                if video_notes:
                    self.excel_writer.WriteDataToExcel(
                        video_notes,
                        USER_VIDEO_FILE_PATH,
                        sheet_name
                    )

                if text_notes:
                    self.excel_writer.WriteDataToExcel(
                        text_notes,
                        USER_TEXT_FILE_PATH,
                        sheet_name
                    )

                print("开始滚动到用户主页底部，以获取已加载笔记总数和最早笔记时间…")
                self.scroll_to_bottom(driver)
                random_delay(2, 1)

                video_count, text_count = self.count_loaded_notes(driver)
                creation_time = self.get_creation_time_from_last_note(driver)

                user_data["Creation Time"] = creation_time
                user_data["Video Count"] = video_count
                user_data["Text Count"] = text_count

                print("-" * 35 + " 最终提取到的用户资料 " + "-" * 35)
                for key, value in user_data.items():
                    print(f"{key}: {value}")
                print("-" * 100)

                self.excel_writer.WriteDataToExcel(
                    [user_data],
                    USER_FILE_PATH,
                    sheet_name
                )

            except Exception as e:
                print(f"访问用户主页 {user_url} 失败：{e}")

        print("所有用户主页访问完毕")

        try:
            driver.get(initial_list_page_url)
            random_delay(3, 2)
        except Exception as e:
            print(f"返回用户搜索结果页失败：{e}")


# =========================
# 8. 主程序
# =========================

def run_integrated_mode(driver):
    """执行综合模式。"""
    spider = RedBookIntegratedSpider()

    search_queries = split_keywords(input("请输入关键词，多个关键词用空格分隔："))

    comment_count_to_crawl = read_non_negative_int(
        "请输入每条笔记需要爬取的评论数量：",
        allow_zero=True
    )

    print("综合模式支持两种输入方式：")
    print("1. 指定总爬取数量，自动统计视频与图文比例")
    print("2. 分别指定视频笔记和图文笔记数量")

    integrate_input_mode = read_choice("请选择综合模式输入方式（1/2）：", {"1", "2"})

    target_total_count = 0
    target_video_count = 0
    target_text_count = 0

    if integrate_input_mode == "1":
        target_total_count = read_non_negative_int(
            "请输入总爬取笔记数量：",
            allow_zero=False
        )
    else:
        target_video_count = read_non_negative_int(
            "请输入要爬取的视频笔记数量：",
            allow_zero=True
        )
        target_text_count = read_non_negative_int(
            "请输入要爬取的图文笔记数量：",
            allow_zero=True
        )

    for query in search_queries:
        print("=" * 100)
        print(f"开始执行综合模式，关键词：{query}")
        print("=" * 100)

        spider.Search(driver, query)
        spider.SelectAllCategory(driver)

        spider.HandleIntegrateMode(
            driver,
            query,
            target_total_count,
            target_video_count,
            target_text_count,
            comment_count_to_crawl
        )

        spider.ClearSearchInput(driver)

    print("综合模式任务执行完毕。")


def run_user_mode(driver):
    """执行用户模式。"""
    spider = RedBookUserSpider()

    search_queries = split_keywords(input("请输入关键词，多个关键词用空格分隔："))

    target_user_count = read_non_negative_int(
        "请输入需要访问的用户数量：",
        allow_zero=False
    )
    target_video_count = read_non_negative_int(
        "请输入每个用户主页要爬取的视频笔记数量：",
        allow_zero=True
    )
    target_text_count = read_non_negative_int(
        "请输入每个用户主页要爬取的图文笔记数量：",
        allow_zero=True
    )
    comment_count_to_crawl = read_non_negative_int(
        "请输入每条笔记需要爬取的评论数量：",
        allow_zero=True
    )

    for query in search_queries:
        print("=" * 100)
        print(f"开始执行用户模式，关键词：{query}")
        print("=" * 100)

        spider.Search(driver, query)
        spider.SelectUserCategory(driver)

        spider.CollectUsersBySearchResults(
            driver,
            query,
            target_user_count,
            target_video_count,
            target_text_count,
            comment_count_to_crawl
        )

        spider.ClearSearchInput(driver)

    print("用户模式任务执行完毕。")


if __name__ == "__main__":
    driver = create_driver()

    try:
        driver.get(REDBOOK_PATH)
        time.sleep(3)

        input("请先完成登录，登录后按回车继续：")

        print("请选择运行模式：")
        print("1. 综合模式：按关键词采集综合搜索结果中的视频/图文笔记")
        print("2. 用户模式：按关键词搜索用户，并采集用户主页资料与笔记")

        run_mode = read_choice("请输入运行模式（1/2）：", {"1", "2"})

        if run_mode == "1":
            run_integrated_mode(driver)
        else:
            run_user_mode(driver)

    finally:
        driver.quit()
